# -*- coding: utf-8 -*-
"""
Created on Wed Sep 13 15:07:33 2023

@author: pan
"""

#导入包
import streamlit as st
import openpyxl
import copy
import pandas as pd
import io
buffer = io.BytesIO()
#print(copy.__version__)
st.set_page_config(
    page_title="quote",    #页面标题
    page_icon=":rainbow:",        #icon:emoji":rainbow:"
    layout="wide",                #页面布局
    initial_sidebar_state="auto"  #侧边栏
)


st.markdown(""" <style> .font {
    font-size:25px ; font-family: 'Cooper Black'; color: #FF9633;}
    </style> """, unsafe_allow_html=True)
st.markdown('<p class="font">请上传您的标签模板，该应用会自动生成标签列表</p>', unsafe_allow_html=True)

#下载导入模板
with open('./标签模板.xlsx', 'rb') as template_file:
    st.download_button(
    label="下载《标签模板》",
    data=template_file,
    file_name='标签模板.xlsx',
    mime='application/vnd.ms-excel'
    )
with st.expander("标签模板使用说明"):
    st.markdown("""
        +  **label**:标签生成结果页（查看即可）
        +  **template**:标签模板（编辑标签样式，支持合并单元格）
        +  **cellset**:单元格设置页（输入标签区域：①起始单元格坐标及结束单元格坐标；②需要合并单元格区域的列表，有多个区域则用逗号分隔）
        +  **sourcedata**:标签数据表（【列名】为该列数据需要填入的【单元格坐标】）
    """) 
    
st.divider()   
 
uploaded_file = st.file_uploader("请选择要上传的xlsx格式标签或拖拽文件至下方区域！",type=['xlsx'])
# row1 = st.columns([1,1,1,1])
# sfe=row1[0].text_input(label="起始单元格字母",key='sfe', value='A').upper()
# sfn=row1[1].number_input(label="起始单元格数字",key='sfn', min_value=1,value=1)
# ste=row1[2].text_input(label="结束单元格字母",key='ste', value='').upper()
# stn=row1[3].number_input(label="结束单元格数字",key='stn', min_value=1)
# hb=st.text_input(label="合并单元格：用英文格式【逗号，】隔开，如：",key='hb', value="C2:D2,E2:F3").upper()
#uploaded_file=r'd:\test.xlsx'


if uploaded_file is not None:

    wb = openpyxl.load_workbook(uploaded_file)

    sourcedata=pd.read_excel(uploaded_file,"sourcedata")
    cellset=pd.read_excel(uploaded_file,"cellset")
    sfe=cellset["起始单元格字母"][0].upper()
    sfn=pd.to_numeric(cellset["起始单元格数字"][0])
    ste=cellset["结束单元格字母"][0].upper()
    stn=pd.to_numeric(cellset["结束单元格数字"][0])
    
    #st.dataframe(sourcedata)
    #st.write(sfe)
    #将模板粘贴到label表
    sws=wb['template']
    tws=wb['label']
    
    #以字符串输入复制、粘贴的区域，如'a1:f16','h23:m38'(必须大小一致)
    #sfe,sfn,ste,stn='a',1,'f',5
    Source_Area =  sfe+str(sfn)+":"+ste+str(stn)
    
    #分别指定复制和粘贴所在sheet的位置（本文复制粘贴的单元格区域都在ws内，ws是什么在上面已经指定好）
    source_area = wb['template'][Source_Area]   
    
    #创造source_cell_list，用以和target_cell_list一一对应：
    source_cell_list = []
    for source_row in source_area:
        for source_cell in source_row:
            sc_str = str(source_cell)  
            point_time = sc_str.count('.')
            sc_str = sc_str.replace('.', '', point_time - 1)
            start = sc_str.find('.')
            sc_str = sc_str[start+1 : -1]
            source_cell_list.append(sc_str) #提取出单元格编号的字符串，如'C8'
    #print('source_cell_list:',source_cell_list)
    
    num=sourcedata.shape[0]
    for j in range(0,num):
        #将数据写入模板
        for h in range(0,sourcedata.shape[1]):
            #i=0
            cellname=sourcedata.columns[h]
            #print(cellname)
            wb['template'][cellname]=sourcedata.iloc[j,h]
    
        #j=0
        tfe=sfe
        tfn=(stn-sfn+1)*j+1
        tte=ste
        ttn=(stn-sfn+1)*(j+1)
        Target_Area = tfe+str(tfn)+":"+tte+str(ttn)
        
        target_area = wb['label'][Target_Area]  
        target_cell_list = []
        for target_row in target_area:
            for target_cell in target_row:
                tc_str = str(target_cell)  
                point_time = tc_str.count('.')
                tc_str = tc_str.replace('.', '', point_time - 1)
                start = tc_str.find('.')
                tc_str = tc_str[start + 1: -1]
                target_cell_list.append(tc_str)  # 提取出单元格编号的字符串，如'L10'
        #print('target_cell_list:',target_cell_list)
         
         
        #获取要复制的单元格总个数：
        cells = len(source_cell_list) 
         
        #提取并复制格式：
        i=0
        while i<=cells-1:
            
            tc=target_cell_list[0+i]#cellname目标单元格
            sc=source_cell_list[0+i]#源单元格
            
            tws[tc].data_type = sws[sc].data_type
            if sws[sc].has_style:
                tws[tc]._style = copy.copy(sws[sc]._style)
                tws[tc].font = copy.copy(sws[sc].font)
                tws[tc].border = copy.copy(sws[sc].border)
                tws[tc].fill = copy.copy(sws[sc].fill)
                tws[tc].number_format = copy.copy(sws[sc].number_format)
                tws[tc].protection = copy.copy(sws[sc].protection)
                tws[tc].alignment = copy.copy(sws[sc].alignment)
                
                #提取数字
                tcn=pd.to_numeric("".join([i for i in tc if i.isdigit()]))
                scn=pd.to_numeric("".join([i for i in sc if i.isdigit()]))
                tws.row_dimensions[tcn].height = sws.row_dimensions[scn].height
                #提取英文
                tce="".join([i for i in tc if i.isalpha()])
                sce="".join([i for i in sc if i.isalpha()])
                tws.column_dimensions[tce].width = sws.column_dimensions[sce].width
                
            # 通过引用方法粘贴值: ws['']=ws[''].value
            tws[tc] = sws[sc].value
            #tws.merge_cells('e3:f3')
            i+=1
    #先填写数据，再合并单元格，合并后会保留左上格的样式
    #hb="C2:D2,E2:F3"
    hb=cellset["需要合并的单元格"][0]
    if ((pd.isna(hb))|(hb=="")|(hb==0)):
        pass
    else:
        hb=str(hb).upper()
        hbdf=pd.DataFrame({"mergecell":hb.split(",")})
        #mergecell start english    
        hbdf['mse']=hbdf['mergecell'].apply(lambda x:"".join([i for i in x[:x.find(":")] if i.isalpha()]))
        hbdf['msn']=hbdf['mergecell'].apply(lambda x:pd.to_numeric("".join([i for i in x[:x.find(":")] if i.isdigit()])))
        hbdf['mee']=hbdf['mergecell'].apply(lambda x:"".join([i for i in x[x.find(":"):] if i.isalpha()]))
        hbdf['men']=hbdf['mergecell'].apply(lambda x:pd.to_numeric("".join([i for i in x[x.find(":"):] if i.isdigit()])))
            
        for hbn in range(0,hbdf.shape[0]):
            #hbn=0
            for j in range(0,num):
                #合并单元格开始行
                tfn=(stn-sfn+1)*j+hbdf['msn'][hbn]
                #合并单元格结束行
                ttn=(stn-sfn+1)*j+hbdf['men'][hbn]
                #组合成合并区域
                merge_Area = hbdf['mse'][hbn]+str(tfn)+":"+hbdf['mee'][hbn]+str(ttn)
                #print(Target_Area)
                tws.merge_cells(merge_Area)
    #保存
    wb.save(buffer)
    st.success('转换成功！可点击下方下载按钮查看标签生成结果')
    st.download_button(
        label="点我下载文件",
        data=buffer,
        file_name="标签结果.xlsx",
        key='concatword',
        mime='application/vnd.ms-excel'
        )
        
    
    
